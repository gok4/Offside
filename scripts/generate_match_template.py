#!/usr/bin/env python3
"""
generate_match_template.py

Generates an Excel data-entry template for a single fixture (home team vs away team).
Pulls both teams' rosters from teams.json so you only ever see the ~50 players
relevant to this match, split into Outfield / Goalkeeper sheets per team, plus a
Match Info sheet and a Team Stats sheet.

Teams and players are identified by their exact "name" field (there is no "slug"
field in teams.json) — team names are matched exactly, and player identification is
always scoped within the team roster being loaded.

Usage:
    python scripts/generate_match_template.py \
        --match-id gfl-2026-001 \
        --matchweek 1 \
        --date 2026-08-30 \
        --kickoff 20:00 \
        --venue "Floodlight Arena" \
        --home "Amsterdam Ravens" \
        --away "Steel Hawks" \
        --teams-file data/teams.json \
        --output-dir data/match_templates

Output:
    data/match_templates/<match-id>.xlsx
"""

import argparse
import json
import os
import sys

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# ---------------------------------------------------------------------------
# Column definitions (must match the rulebook's §2.1 / §2.2 data schema)
# ---------------------------------------------------------------------------

OUTFIELD_STAT_COLUMNS = [
    "Minutes Played",
    "Goals",
    "Assists",
    "Accurate Passes (Made)",
    "Accurate Passes (Attempted)",
    "Chances Created",
    "Shots On Target",
    "Shots Off Target",
    "Blocked Shots",
    "Touches",
    "Touches In Opposition Box",
    "Successful Dribbles (Made)",
    "Successful Dribbles (Attempted)",
    "Passes Into Final Third",
    "Dispossessed",
    "Tackles",
    "Blocks",
    "Clearances",
    "Interceptions",
    "Recoveries",
    "Dribbled Past",
    "Ground Duels Won (Won)",
    "Ground Duels Won (Attempted)",
    "Aerial Duels Won (Won)",
    "Aerial Duels Won (Attempted)",
    "Was Fouled",
    "Fouls Committed",
    "Yellow Card (Y/N)",
    "Red Card (Y/N)",
    "Own Goal (Y/N)",
    "Penalty Miss (Y/N)",
]

GK_STAT_COLUMNS = [
    "Minutes Played",
    "Saves",
    "Goals Conceded",
    "Accurate Passes (Made)",
    "Accurate Passes (Attempted)",
    "Accurate Long Balls (Made)",
    "Accurate Long Balls (Attempted)",
    "Diving Save",
    "Saves Inside Box",
    "Acted As Sweeper",
    "Punches",
    "Throws",
    "High Claim",
    "Recoveries",
    "Clearances",
    "Touches",
    "Ground Duels Won (Won)",
    "Ground Duels Won (Attempted)",
    "Penalty Save (Y/N)",
    "Yellow Card (Y/N)",
    "Red Card (Y/N)",
]

PLAYED_STATUS_OPTIONS = ["Starting XI", "Sub (Came On)", "Sub (Unused)", "Did Not Play"]

# Matchday squad rules (Rulebook reference)
REQUIRED_STARTING_XI = 11
REQUIRED_TOTAL_SUBS = 7          # "Sub (Came On)" + "Sub (Unused)" combined
MAX_SUBSTITUTIONS_USED = 4       # max "Sub (Came On)" count

# Styling constants
HEADER_FILL = PatternFill(start_color="0B1F3A", end_color="0B1F3A", fill_type="solid")  # navy
HEADER_FONT = Font(color="D4AF37", bold=True)  # gold on navy, matches site theme
LOCKED_FILL = PatternFill(start_color="E8E8E8", end_color="E8E8E8", fill_type="solid")
TITLE_FONT = Font(bold=True, size=14)


def load_team_roster(teams_file, team_name):
    """Load a single team's player roster from teams.json, matched by exact team name."""
    with open(teams_file, "r", encoding="utf-8") as f:
        data = json.load(f)

    # Support both {"teams": [...]} and a flat list at the top level.
    teams = data.get("teams", data) if isinstance(data, dict) else data

    for team in teams:
        if team.get("name") == team_name:
            return team

    available = ", ".join(t.get("name", "?") for t in teams)
    raise ValueError(
        f"Team '{team_name}' not found in {teams_file}.\nAvailable teams: {available}"
    )


def split_roster(team):
    """Split a team's players into outfield vs goalkeeper lists based on position."""
    outfield, goalkeepers = [], []
    for player in team.get("players", []):
        position = (player.get("position") or "").strip().upper()
        if position in ("GK", "GOALKEEPER"):
            goalkeepers.append(player)
        else:
            outfield.append(player)
    return outfield, goalkeepers


def style_header_row(ws, row_idx, num_cols):
    for col_idx in range(1, num_cols + 1):
        cell = ws.cell(row=row_idx, column=col_idx)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def autosize_columns(ws, min_width=10, max_width=32):
    for col_cells in ws.columns:
        length = max((len(str(c.value)) for c in col_cells if c.value is not None), default=0)
        col_letter = get_column_letter(col_cells[0].column)
        ws.column_dimensions[col_letter].width = max(min_width, min(max_width, length + 2))


def build_roster_sheet(wb, sheet_name, players, stat_columns):
    ws = wb.create_sheet(sheet_name)

    # Locked reference column (Player Name) first, then "Played?" status, then stats.
    headers = ["Player Name", "Position", "Played?"] + stat_columns
    ws.append(headers)
    style_header_row(ws, 1, len(headers))
    ws.freeze_panes = "D2"  # keep name/position/played columns visible while scrolling stats

    played_col_letter = get_column_letter(3)  # "Played?" is column C

    for row_idx, player in enumerate(players, start=2):
        name = player.get("name", "")
        position = player.get("position", "")
        ws.cell(row=row_idx, column=1, value=name).fill = LOCKED_FILL
        ws.cell(row=row_idx, column=2, value=position).fill = LOCKED_FILL
        # Leave "Played?" and all stat columns blank for manual entry.

    # Dropdown validation for the "Played?" column.
    dv = DataValidation(
        type="list",
        formula1=f'"{",".join(PLAYED_STATUS_OPTIONS)}"',
        allow_blank=True,
    )
    ws.add_data_validation(dv)
    last_row = len(players) + 1
    dv.add(f"{played_col_letter}2:{played_col_letter}{last_row}")

    autosize_columns(ws)
    return ws


def build_match_info_sheet(wb, args):
    ws = wb.create_sheet("Match Info", 0)  # first sheet
    ws.append(["Field", "Value"])
    style_header_row(ws, 1, 2)

    rows = [
        ("Match ID", args.match_id),
        ("Matchweek", args.matchweek),
        ("Date", args.date),
        ("Kickoff", args.kickoff or ""),
        ("Venue", args.venue or ""),
        ("Home Team", args.home),
        ("Away Team", args.away),
        ("Home Team Score", ""),   # fill in after the match
        ("Away Team Score", ""),   # fill in after the match
        ("Status", "completed"),  # scheduled / live / completed
    ]
    for r in rows:
        ws.append(r)

    autosize_columns(ws)
    return ws


def build_team_stats_sheet(wb, args):
    ws = wb.create_sheet("Team Stats")
    headers = ["Stat", "Home Team", "Away Team"]
    ws.append(headers)
    style_header_row(ws, 1, len(headers))

    stat_rows = [
        "Total Shots",
        "Shots On Target",
        "Touches In Opposition Box",
        "Accurate Passes",
        "Yellow Cards",
    ]
    for stat in stat_rows:
        ws.append([stat, "", ""])

    autosize_columns(ws)
    return ws


def load_all_team_names(teams_file):
    """Return the full list of team names in teams.json, in file order."""
    with open(teams_file, "r", encoding="utf-8") as f:
        data = json.load(f)
    teams = data.get("teams", data) if isinstance(data, dict) else data
    return [t.get("name") for t in teams]


def pick_team_interactively(team_names, role_label):
    """Print a numbered list of teams and prompt the user to pick one."""
    print(f"\nSelect the {role_label} team:")
    for i, name in enumerate(team_names, start=1):
        print(f"  {i:>2}. {name}")
    while True:
        choice = input(f"{role_label} team number: ").strip()
        if choice.isdigit() and 1 <= int(choice) <= len(team_names):
            return team_names[int(choice) - 1]
        print("Invalid selection, try again.")


def prompt_if_missing(value, prompt_text, required=True, cast=str):
    if value not in (None, ""):
        return value
    while True:
        raw = input(prompt_text).strip()
        if raw or not required:
            try:
                return cast(raw) if raw else raw
            except ValueError:
                print("Invalid value, try again.")
                continue
        print("This field is required.")


def main():
    parser = argparse.ArgumentParser(
        description="Generate a match data-entry Excel template. "
        "Run with no arguments for an interactive team picker."
    )
    parser.add_argument("--match-id")
    parser.add_argument("--matchweek", type=int)
    parser.add_argument("--date", help="YYYY-MM-DD")
    parser.add_argument("--kickoff", default="", help="HH:MM (optional)")
    parser.add_argument("--venue", default="", help="Venue name (optional)")
    parser.add_argument("--home", help="Home team name (exact match, e.g. 'Amsterdam Ravens')")
    parser.add_argument("--away", help="Away team name (exact match, e.g. 'Steel Hawks')")
    parser.add_argument("--teams-file", default="data/teams.json")
    parser.add_argument("--output-dir", default="data/match_templates")
    args = parser.parse_args()

    if not os.path.isfile(args.teams_file):
        print(f"ERROR: teams file not found at '{args.teams_file}'", file=sys.stderr)
        sys.exit(1)

    team_names = load_all_team_names(args.teams_file)

    # Interactive fallback: if --home/--away weren't passed, show a numbered picker.
    if not args.home:
        args.home = pick_team_interactively(team_names, "HOME")
    if not args.away:
        args.away = pick_team_interactively(team_names, "AWAY")

    args.match_id = prompt_if_missing(args.match_id, "Match ID (e.g. gfl-2026-001): ")
    args.matchweek = prompt_if_missing(args.matchweek, "Matchweek (number): ", cast=int)
    args.date = prompt_if_missing(args.date, "Date (YYYY-MM-DD): ")
    if not args.kickoff:
        args.kickoff = input("Kickoff (HH:MM, optional): ").strip()
    if not args.venue:
        args.venue = input("Venue (optional): ").strip()

    home_team = load_team_roster(args.teams_file, args.home)
    away_team = load_team_roster(args.teams_file, args.away)

    home_outfield, home_gk = split_roster(home_team)
    away_outfield, away_gk = split_roster(away_team)

    wb = Workbook()
    wb.remove(wb.active)  # drop the default blank sheet

    build_match_info_sheet(wb, args)
    build_roster_sheet(wb, "Home - Outfield", home_outfield, OUTFIELD_STAT_COLUMNS)
    build_roster_sheet(wb, "Home - GK", home_gk, GK_STAT_COLUMNS)
    build_roster_sheet(wb, "Away - Outfield", away_outfield, OUTFIELD_STAT_COLUMNS)
    build_roster_sheet(wb, "Away - GK", away_gk, GK_STAT_COLUMNS)
    build_team_stats_sheet(wb, args)

    os.makedirs(args.output_dir, exist_ok=True)
    output_path = os.path.join(args.output_dir, f"{args.match_id}.xlsx")
    wb.save(output_path)

    print(f"\nTemplate created: {output_path}")
    print(f"  Home ({args.home}): {len(home_outfield)} outfield, {len(home_gk)} GK")
    print(f"  Away ({args.away}): {len(away_outfield)} outfield, {len(away_gk)} GK")
    print(
        f"Matchday squad rules: {REQUIRED_STARTING_XI} Starting XI + "
        f"{REQUIRED_TOTAL_SUBS} substitutes (max {MAX_SUBSTITUTIONS_USED} used) per team."
    )
    print("Fill in 'Played?', stats for players who featured, and the Team Stats sheet.")


if __name__ == "__main__":
    main()

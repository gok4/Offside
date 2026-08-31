#!/usr/bin/env python3
"""
calculate_match_points.py

Reads a filled-in match Excel template (produced by generate_match_template.py),
looks up each player's position from teams.json, applies the Offside GFL scoring
rulebook (fantasy points + ICT Index), and writes the final match JSON to
data/matches/<match-id>.json.

Usage:
    python scripts/calculate_match_points.py \
        --input data/match_templates/gfl-2026-001.xlsx \
        --teams-file data/teams.json \
        --output-dir data/matches

Output:
    data/matches/<match-id>.json
"""

import argparse
import json
import os
import sys

from openpyxl import load_workbook

# ---------------------------------------------------------------------------
# Position bucket mapping — matches main.js's POSITION_GROUPS exactly.
# ---------------------------------------------------------------------------

POSITION_TO_BUCKET = {
    "GK": "GK",
    "RB": "DEF", "CB": "DEF", "LB": "DEF",
    "CDM": "MID", "CM": "MID", "CAM": "MID", "RM": "MID", "LM": "MID",
    "RW": "FWD", "LW": "FWD", "ST": "FWD",
}

# Matchday squad rules (must match generate_match_template.py)
REQUIRED_STARTING_XI = 11
REQUIRED_TOTAL_SUBS = 7          # "Sub (Came On)" + "Sub (Unused)" combined
MAX_SUBSTITUTIONS_USED = 4       # max "Sub (Came On)" count

# ---------------------------------------------------------------------------
# Scoring tables (Rulebook §3)
# ---------------------------------------------------------------------------

GOAL_POINTS = {"GK": 10, "DEF": 6, "MID": 5, "FWD": 4}
ASSIST_POINTS = 3
CLEAN_SHEET_POINTS = {"GK": 4, "DEF": 4, "MID": 1, "FWD": 0}

YELLOW_CARD_POINTS = -1
RED_CARD_POINTS = -3  # replaces yellow deduction, not additive
OWN_GOAL_POINTS = -2
PENALTY_MISS_POINTS = -2
PENALTY_SAVE_POINTS = 5

DEF_CBIT_THRESHOLD = 10        # Defenders: Clearances+Blocks+Interceptions+Tackles
MID_FWD_CBIT_THRESHOLD = 12    # Mid/Fwd: + Recoveries
DEFENSIVE_CONTRIBUTION_POINTS = 2

SHOT_SAVES_PER_POINT = 3       # every 3 shot saves = 1 point
GOALS_CONCEDED_PER_DEDUCTION = 2  # every 2 goals conceded (GK/DEF) = -1


def truthy(value):
    """Interpret a Y/N-style Excel cell as a boolean."""
    if value is None:
        return False
    text = str(value).strip().lower()
    return text in ("y", "yes", "true", "1")


def load_team_lookup(teams_file):
    """Build {team_name: {player_name: position}} from teams.json."""
    with open(teams_file, "r", encoding="utf-8") as f:
        data = json.load(f)
    teams = data.get("teams", data) if isinstance(data, dict) else data

    lookup = {}
    for team in teams:
        team_name = team.get("name")
        players = {p.get("name"): p.get("position") for p in team.get("players", [])}
        lookup[team_name] = players
    return lookup


def read_match_info(wb):
    ws = wb["Match Info"]
    info = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        field, value = row[0], row[1]
        if field:
            info[field] = value
    return info


def read_team_stats(wb):
    ws = wb["Team Stats"]
    home_stats, away_stats = {}, {}
    key_map = {
        "Total Shots": "totalShots",
        "Shots On Target": "shotsOnTarget",
        "Touches In Opposition Box": "touchesInOppositionBox",
        "Accurate Passes": "accuratePasses",
        "Yellow Cards": "yellowCards",
    }
    for row in ws.iter_rows(min_row=2, values_only=True):
        stat_label, home_val, away_val = row[0], row[1], row[2]
        if stat_label in key_map:
            home_stats[key_map[stat_label]] = home_val if home_val is not None else 0
            away_stats[key_map[stat_label]] = away_val if away_val is not None else 0
    return home_stats, away_stats


def read_roster_sheet(wb, sheet_name, is_gk):
    """Read a roster sheet, returning:
      - results: {player_name: {played_status, stats}} for players who took the field
      - unused_subs: [player_name, ...] named on the bench but never came on
    """
    ws = wb[sheet_name]
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    header_index = {h: i for i, h in enumerate(headers)}

    results = {}
    unused_subs = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        name = row[header_index["Player Name"]]
        played_status = row[header_index["Played?"]]
        if not name or not played_status or played_status in ("Did Not Play", "Sub (Unused)"):
            if name and played_status == "Sub (Unused)":
                unused_subs.append(name)
            continue

        def get(col_name, default=0):
            idx = header_index.get(col_name)
            if idx is None:
                return default
            val = row[idx]
            return val if val is not None else default

        if is_gk:
            stats = {
                "minutesPlayed": get("Minutes Played"),
                "saves": get("Saves"),
                "goalsConceded": get("Goals Conceded"),
                "accuratePasses": {"made": get("Accurate Passes (Made)"), "attempted": get("Accurate Passes (Attempted)")},
                "accurateLongBalls": {"made": get("Accurate Long Balls (Made)"), "attempted": get("Accurate Long Balls (Attempted)")},
                "divingSave": get("Diving Save"),
                "savesInsideBox": get("Saves Inside Box"),
                "actedAsSweeper": get("Acted As Sweeper"),
                "punches": get("Punches"),
                "throws": get("Throws"),
                "highClaim": get("High Claim"),
                "recoveries": get("Recoveries"),
                "clearances": get("Clearances"),
                "touches": get("Touches"),
                "groundDuelsWon": {"won": get("Ground Duels Won (Won)"), "attempted": get("Ground Duels Won (Attempted)")},
                "penaltySave": truthy(get("Penalty Save (Y/N)", "N")),
                "yellowCard": truthy(get("Yellow Card (Y/N)", "N")),
                "redCard": truthy(get("Red Card (Y/N)", "N")),
            }
        else:
            stats = {
                "minutesPlayed": get("Minutes Played"),
                "goals": get("Goals"),
                "assists": get("Assists"),
                "accuratePasses": {"made": get("Accurate Passes (Made)"), "attempted": get("Accurate Passes (Attempted)")},
                "chancesCreated": get("Chances Created"),
                "shotsOnTarget": get("Shots On Target"),
                "shotsOffTarget": get("Shots Off Target"),
                "blockedShots": get("Blocked Shots"),
                "touches": get("Touches"),
                "touchesInOppositionBox": get("Touches In Opposition Box"),
                "successfulDribbles": {"made": get("Successful Dribbles (Made)"), "attempted": get("Successful Dribbles (Attempted)")},
                "passesIntoFinalThird": get("Passes Into Final Third"),
                "dispossessed": get("Dispossessed"),
                "tackles": get("Tackles"),
                "blocks": get("Blocks"),
                "clearances": get("Clearances"),
                "interceptions": get("Interceptions"),
                "recoveries": get("Recoveries"),
                "dribbledPast": get("Dribbled Past"),
                "groundDuelsWon": {"won": get("Ground Duels Won (Won)"), "attempted": get("Ground Duels Won (Attempted)")},
                "aerialDuelsWon": {"won": get("Aerial Duels Won (Won)"), "attempted": get("Aerial Duels Won (Attempted)")},
                "wasFouled": get("Was Fouled"),
                "foulsCommitted": get("Fouls Committed"),
                "yellowCard": truthy(get("Yellow Card (Y/N)", "N")),
                "redCard": truthy(get("Red Card (Y/N)", "N")),
                "ownGoal": truthy(get("Own Goal (Y/N)", "N")),
                "penaltyMiss": truthy(get("Penalty Miss (Y/N)", "N")),
            }

        results[name] = {"played_status": played_status, "stats": stats}
    return results, unused_subs


def validate_squad(team_name, starting_xi, substitutes_came_on, unused_subs):
    """Check the matchday squad against the 11 / 7 / max-4-subs rules. Returns a list of warnings."""
    warnings = []
    if len(starting_xi) != REQUIRED_STARTING_XI:
        warnings.append(
            f"{team_name}: expected {REQUIRED_STARTING_XI} in Starting XI, found {len(starting_xi)}."
        )
    total_subs = len(substitutes_came_on) + len(unused_subs)
    if total_subs != REQUIRED_TOTAL_SUBS:
        warnings.append(
            f"{team_name}: expected {REQUIRED_TOTAL_SUBS} named substitutes (used + unused), "
            f"found {total_subs}."
        )
    if len(substitutes_came_on) > MAX_SUBSTITUTIONS_USED:
        warnings.append(
            f"{team_name}: {len(substitutes_came_on)} substitutions used, "
            f"exceeds the max of {MAX_SUBSTITUTIONS_USED}."
        )
    return warnings


# ---------------------------------------------------------------------------
# Fantasy points calculation (Rulebook §3)
# ---------------------------------------------------------------------------

def calculate_points(stats, bucket, is_gk, team_conceded, opponent_played_60):
    """Calculate fantasy points for a single player for this match."""
    points = 0
    breakdown = {}

    def add(label, value):
        nonlocal points
        if value:
            breakdown[label] = value
            points += value

    minutes = stats.get("minutesPlayed", 0) or 0

    # Appearance
    if minutes >= 60:
        add("appearance", 2)
    elif minutes > 0:
        add("appearance", 1)

    if is_gk:
        add("goals", stats.get("goals", 0) * GOAL_POINTS["GK"] if stats.get("goals") else 0)
    else:
        goals = stats.get("goals", 0) or 0
        add("goals", goals * GOAL_POINTS.get(bucket, 0))

    assists = stats.get("assists", 0) or 0
    add("assists", assists * ASSIST_POINTS)

    # Clean sheet: only if played 60+ minutes and team didn't concede.
    if minutes >= 60 and team_conceded == 0:
        add("clean_sheet", CLEAN_SHEET_POINTS.get(bucket, 0))

    if is_gk:
        saves = stats.get("saves", 0) or 0
        add("shot_saves", (saves // SHOT_SAVES_PER_POINT) * 1)

        if stats.get("penaltySave"):
            add("penalty_save", PENALTY_SAVE_POINTS)

    # Goals conceded deduction (GK/DEF only) — every 2 goals conceded = -1 point.
    # For GKs this uses their own "Goals Conceded" entry. For outfield defenders, there is
    # no per-player "goals conceded" stat in the data feed, so this uses the team's total
    # goals conceded for the match as an approximation (does not account for a defender
    # being substituted before a late goal was conceded).
    if bucket == "GK":
        conceded = stats.get("goalsConceded", 0) or 0
        add("goals_conceded_deduction", -(conceded // GOALS_CONCEDED_PER_DEDUCTION))
    elif bucket == "DEF":
        conceded = team_conceded or 0
        add("goals_conceded_deduction", -(conceded // GOALS_CONCEDED_PER_DEDUCTION))

    # Defensive contributions (threshold-based, non-stacking)
    if not is_gk:
        cbit = (
            (stats.get("clearances", 0) or 0)
            + (stats.get("blocks", 0) or 0) + (stats.get("blockedShots", 0) or 0)
            + (stats.get("interceptions", 0) or 0)
            + (stats.get("tackles", 0) or 0)
        )
        if bucket == "DEF" and cbit >= DEF_CBIT_THRESHOLD:
            add("defensive_contribution", DEFENSIVE_CONTRIBUTION_POINTS)
        elif bucket in ("MID", "FWD"):
            cbit_plus_recoveries = cbit + (stats.get("recoveries", 0) or 0)
            if cbit_plus_recoveries >= MID_FWD_CBIT_THRESHOLD:
                add("defensive_contribution", DEFENSIVE_CONTRIBUTION_POINTS)

    # Discipline
    if stats.get("redCard"):
        add("red_card", RED_CARD_POINTS)
    elif stats.get("yellowCard"):
        add("yellow_card", YELLOW_CARD_POINTS)

    if stats.get("ownGoal"):
        add("own_goal", OWN_GOAL_POINTS)

    if stats.get("penaltyMiss"):
        add("penalty_miss", PENALTY_MISS_POINTS)

    return points, breakdown


# ---------------------------------------------------------------------------
# Offside ICT Index (Rulebook §4)
# ---------------------------------------------------------------------------

def calculate_ict(stats, is_gk):
    if is_gk:
        influence = (
            2 * (stats.get("saves", 0) or 0)
            + 1 * (stats.get("divingSave", 0) or 0)
            + 1 * (stats.get("savesInsideBox", 0) or 0)
            + 1 * (stats.get("actedAsSweeper", 0) or 0)
            + 1 * (stats.get("highClaim", 0) or 0)
            + 0.5 * (stats.get("punches", 0) or 0)
            + 1 * (stats.get("groundDuelsWon", {}).get("won", 0) or 0)
            - 3 * (stats.get("goalsConceded", 0) or 0)
        )
        creativity = (
            1 * (stats.get("accurateLongBalls", {}).get("made", 0) or 0)
            + 0.1 * (stats.get("accuratePasses", {}).get("made", 0) or 0)
            + 0.2 * (stats.get("throws", 0) or 0)
        )
        threat = 0
    else:
        influence = (
            10 * (stats.get("goals", 0) or 0)
            + 6 * (stats.get("assists", 0) or 0)
            + 2 * (stats.get("tackles", 0) or 0)
            + 2 * (stats.get("interceptions", 0) or 0)
            + 2 * (stats.get("blocks", 0) or 0)
            + 1 * (stats.get("clearances", 0) or 0)
            + 1 * (stats.get("recoveries", 0) or 0)
            + 1 * (stats.get("groundDuelsWon", {}).get("won", 0) or 0)
            + 1 * (stats.get("aerialDuelsWon", {}).get("won", 0) or 0)
            + 0.5 * (stats.get("wasFouled", 0) or 0)
            - 0.5 * (stats.get("foulsCommitted", 0) or 0)
            - 0.5 * (stats.get("dispossessed", 0) or 0)
            - 1 * (stats.get("dribbledPast", 0) or 0)
        )
        creativity = (
            3 * (stats.get("chancesCreated", 0) or 0)
            + 2 * (stats.get("assists", 0) or 0)
            + 1 * (stats.get("passesIntoFinalThird", 0) or 0)
            + 1.5 * (stats.get("successfulDribbles", {}).get("made", 0) or 0)
            + 1 * (stats.get("touchesInOppositionBox", 0) or 0)
            + 0.1 * (stats.get("accuratePasses", {}).get("made", 0) or 0)
        )
        threat = (
            5 * (stats.get("goals", 0) or 0)
            + 2 * (stats.get("shotsOnTarget", 0) or 0)
            + 1 * (stats.get("shotsOffTarget", 0) or 0)
            + 1 * (stats.get("blockedShots", 0) or 0)
            + 1 * (stats.get("successfulDribbles", {}).get("made", 0) or 0)
            + 1 * (stats.get("touchesInOppositionBox", 0) or 0)
        )

    ict_index = (influence + creativity + threat) / 10
    return {
        "influence": round(influence, 2),
        "creativity": round(creativity, 2),
        "threat": round(threat, 2),
        "ictIndex": round(ict_index, 2),
    }


def process_team(sheet_prefix, wb, team_lookup, team_name, opponent_conceded_by_team):
    """Process one team's Outfield + GK sheets into a playerStats dict with points/ICT.

    Returns: (player_stats_out, starting_xi, substitutes_came_on, unused_subs, validation_warnings)
    """
    outfield, outfield_unused = read_roster_sheet(wb, f"{sheet_prefix} - Outfield", is_gk=False)
    goalkeepers, gk_unused = read_roster_sheet(wb, f"{sheet_prefix} - GK", is_gk=True)
    unused_subs = outfield_unused + gk_unused

    roster_positions = team_lookup.get(team_name, {})
    player_stats_out = {}
    starting_xi, substitutes_came_on = [], []

    for name, entry in {**outfield, **goalkeepers}.items():
        is_gk = name in goalkeepers
        position = roster_positions.get(name)
        if position is None:
            print(f"WARNING: '{name}' not found in {team_name}'s roster in teams.json — skipping.", file=sys.stderr)
            continue
        bucket = POSITION_TO_BUCKET.get(position, "MID")

        stats = entry["stats"]
        points, breakdown = calculate_points(
            stats, bucket, is_gk,
            team_conceded=opponent_conceded_by_team,
            opponent_played_60=None,
        )
        ict = calculate_ict(stats, is_gk)

        stats["points"] = points
        stats["pointsBreakdown"] = breakdown
        stats["ict"] = ict
        player_stats_out[name] = stats

        if entry["played_status"] == "Starting XI":
            starting_xi.append(name)
        elif entry["played_status"] == "Sub (Came On)":
            substitutes_came_on.append(name)

    warnings = validate_squad(team_name, starting_xi, substitutes_came_on, unused_subs)
    return player_stats_out, starting_xi, substitutes_came_on, unused_subs, warnings


def main():
    parser = argparse.ArgumentParser(description="Calculate fantasy points from a filled-in match template.")
    parser.add_argument("--input", required=True, help="Path to the filled-in .xlsx match template")
    parser.add_argument("--teams-file", default="data/teams.json")
    parser.add_argument("--output-dir", default="data/matches")
    args = parser.parse_args()

    if not os.path.isfile(args.input):
        print(f"ERROR: input file not found at '{args.input}'", file=sys.stderr)
        sys.exit(1)

    wb = load_workbook(args.input, data_only=True)
    team_lookup = load_team_lookup(args.teams_file)

    info = read_match_info(wb)
    home_team_stats, away_team_stats = read_team_stats(wb)

    home_score = info.get("Home Team Score") or 0
    away_score = info.get("Away Team Score") or 0

    home_players, home_xi, home_subs, home_unused, home_warnings = process_team(
        "Home", wb, team_lookup, info.get("Home Team"), opponent_conceded_by_team=away_score
    )
    away_players, away_xi, away_subs, away_unused, away_warnings = process_team(
        "Away", wb, team_lookup, info.get("Away Team"), opponent_conceded_by_team=home_score
    )

    all_warnings = home_warnings + away_warnings
    if all_warnings:
        print("\n--- SQUAD VALIDATION WARNINGS ---", file=sys.stderr)
        for w in all_warnings:
            print(f"  ! {w}", file=sys.stderr)
        print("Continuing anyway — review the template if these look wrong.\n", file=sys.stderr)

    match_json = {
        "matchId": info.get("Match ID"),
        "matchweek": info.get("Matchweek"),
        "date": str(info.get("Date")),
        "kickoff": info.get("Kickoff"),
        "venue": info.get("Venue"),
        "status": info.get("Status", "completed"),
        "homeTeam": {
            "name": info.get("Home Team"),
            "score": home_score,
            "startingXI": home_xi,
            "substitutes": home_subs,
            "unusedSubstitutes": home_unused,
            "teamStats": home_team_stats,
            "playerStats": home_players,
        },
        "awayTeam": {
            "name": info.get("Away Team"),
            "score": away_score,
            "startingXI": away_xi,
            "substitutes": away_subs,
            "unusedSubstitutes": away_unused,
            "teamStats": away_team_stats,
            "playerStats": away_players,
        },
    }

    os.makedirs(args.output_dir, exist_ok=True)
    match_id = info.get("Match ID") or "unknown-match"
    output_path = os.path.join(args.output_dir, f"{match_id}.json")
    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(match_json, f, indent=2, ensure_ascii=False)

    print(f"Match points calculated: {output_path}")
    print(f"  {info.get('Home Team')} {home_score} - {away_score} {info.get('Away Team')}")
    total_home_pts = sum(p["points"] for p in home_players.values())
    total_away_pts = sum(p["points"] for p in away_players.values())
    print(f"  Total fantasy points — Home: {total_home_pts}, Away: {total_away_pts}")


if __name__ == "__main__":
    main()
